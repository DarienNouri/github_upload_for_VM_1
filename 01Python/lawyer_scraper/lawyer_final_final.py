from requests_html import HTMLSession
import re
from bs4 import BeautifulSoup
import pandas as pd
import csv
import time
from bs4 import BeautifulSoup
import urllib3
from seleniumwire import webdriver
import lxml.html as html
import lxml.html.clean as clean
from seleniumwire import webdriver
import requests
import lxml
from lxml import etree
import xml.etree.ElementTree as ET 
from googlesearch import search
from pyvirtualdisplay import Display
from selenium.webdriver import DesiredCapabilities
import pandas as pd
from selenium.webdriver.common.by import By
import openpyxl
from googlesearch import search

x=0
def pull_name():
    name = ''
    name+=(text[0])
    return name

def pull_addy():
    mail_addy = []
    for i in range(len(text)):
        x = 1
        if 'Mail Address:' in text[i]:
            mail_addy.append(text[i+x])
            i +=x
            while ':' not in text[i+x]:
                mail_addy.append(text[i+x])
                i+=x
    mail_addy = mail_addy[-2]
    return mail_addy
def pull_email():
    email = []
    for i in range(len(text)):
        x=1
        if 'Email:' in text[i]:
            email.append(text[i+x])
            i +=x
            while ':' not in text[i+x]:
                email.append(text[i+x])
                i+=x
    return email

def pull_office():
    office = []
    for i in range(len(text)):
          x=0
          if 'Office:' in text[i]:
            office.append(text[i+x])
            i +=x
            while ':' not in text[i+x]:
                office.append(text[i+x])
                i+=x
    return office

def pull_cell():
    cell=[]
    for i in range(len(text)):
          x=0
          if 'Cell:' in text[i]:
                 
            cell.append(text[i+x])
            i +=x
            while ':' not in text[i+x]:
                cell.append(text[i+x])
                i+=x
    return cell
def pull_admitted():
    admitted=[]
    for i in range(len(text)):
          x=1
          if 'Admitted:' in text[i]:
            admitted.append(text[i+x])
            i +=x
            while ':' not in text[i+x]:
                admitted.append(text[i+x])
                i+=x
    return admitted

def pull_firm():
    firm=[]
    for i in range(len(text)):
          x=1
          if 'Firm:' in text[i]:
            firm.append(text[i+x])
            i +=x
            
    return firm

def pull_firm_pos():
    firm_pos=[]
    for i in range(len(text)):
          x=1
          if 'Firm Position:' in text[i]:
            firm_pos.append(text[i+x])
            i +=x
           
    return firm_pos
def pull_firm_site():
    site=[]
    for i in range(len(text)):
          x=1
          
    return site
save = 0
loop=1
df = pd.DataFrame()
data = pd.read_excel(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\lawyer_input.xlsx")
print (data)



wrkbk = openpyxl.load_workbook(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\lawyer_input.xlsx")
sh = wrkbk.active
for i in range(4,sh.max_row+1): #sh.max_row+1
    print("\n")
    
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    link_counter = 0
    
    for link in search(query, tld="co.in", num =1, stop=9, pause =9):  #chage stop= to change how many links are searched
        time.sleep(3)
        link_counter+=1
        session = HTMLSession()
        req_headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.8',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/4.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36'
        }
        try:
            r = session.get(link,  headers=req_headers)
        except:
            print("badlink")
            continue
        if r.status_code == 429:
             time.sleep(4)
        whitelist_url = ['floridabar']
        if any(ext in link for ext in whitelist_url): 
            options = webdriver.ChromeOptions()

            def interceptor(request):
                del request.headers['Referer']  # Delete the header first
                request.headers['Referer'] = 'some_referer'
                options.add_experimental_option('excludeSwitches', ['enable-logging'])
            driver = webdriver.Chrome(options=options)
            # Set the interceptor on the driver
            driver.request_interceptor = interceptor
            # All requests will now use 'some_referer' for the referer
            try:
                driver.get(link)
            except:
                break
            profile = driver.find_element(By.ID, 'mProfile').text
            text = profile.split('\n')
            name = pull_name()
            email =pull_email()
            office = pull_office()
            cell = pull_cell()
            admitted = pull_admitted()
            firm = pull_firm()
            firm_pos = pull_firm_pos()
            firm_size = pull_firm_site()
            addy = pull_addy()          
            
            new_row ={loop:[name,email,office,cell,admitted,firm,firm_pos,firm_size,addy]}
            newnew = {'Name':name,'Email':email,'Office':office, 'Cell':cell, 'Passed Bar Date':admitted, 'Firm':firm, 'Firm Position':firm_pos, 'Firm Size':firm_size,'Address':addy}
            df = df.append(newnew,ignore_index=True)
            loop+=1
            save+=1
            print(df)
            if save >=1:
            
               
                df.to_csv('attorney_list.csv', encoding='utf-8', index=False)
                
                save = 0


            break
        
        
        
        
        
        
        
        
        else:
            continue