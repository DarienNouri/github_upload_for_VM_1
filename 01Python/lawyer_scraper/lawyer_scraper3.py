
from requests_html import HTMLSession
import re
from bs4 import BeautifulSoup
import pandas as pd
import csv
import time

link = 'https://www.floridabar.org/directories/find-mbr/profile/?num=561347'

from bs4 import BeautifulSoup
import urllib3
from seleniumwire import webdriver

moon_url = 'https://www.floridabar.org/directories/find-mbr/profile/?num=564737'


import lxml.html as html
import lxml.html.clean as clean
from seleniumwire import webdriver
import requests


from pyvirtualdisplay import Display
from selenium.webdriver import DesiredCapabilities


import openpyxl


data = pd.read_excel(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\loopnet1test.xlsx")
print (data)
counter = 2
df = pd.DataFrame(data, columns= ['name_email'])
savecount = 0
saveloop = 0


final_list =[]
final = {0: ["emails", "query", "link"]}
wrkbk = openpyxl.load_workbook(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\loopnet1test.xlsx")
sh = wrkbk.active
for i in range(2,sh.max_row+1): #sh.max_row+1
    print("\n")
    print(i)
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    link_counter = 0
    savecount+=1
    for link in search(query, tld="co.in", num =1, stop=8, pause =8):  #chage stop= to change how many links are searched
        
        
        link_counter+=1
        print(query)
        session = HTMLSession()
        
        req_headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.8',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/4.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36'
        }
        
        try:
            r = session.get(link,  headers=req_headers)
        except:
            print("badlink")
            continue
        if r.status_code == 429:
             time.sleep(10)
        time.sleep(.8)
        '''
        def open_html(path):
            with open(path, 'rb') as f:
                return f.read()
            
            
        html = open_html(str(link))
        print(html)
        if r.status_code == 429:
             time.sleep(10)
        time.sleep(.8)
        '''
        #emails = set()  
      
        blacklist_link = ["rocket","linkedin","streeteasy","crexi","zoom","loopnet","zillow"]
        blacklist_email = ["info", "customer", "streeteasy", "help","edu",".org","Info","_","fake","foo", "help","contact","jane","doe",".ru",".io"]
        whitelist_url = ['floridabar']
        

        if any(ext in link for ext in whitelist_url): 
            address = []
            phone = []
            area = []
            firm = []
            law_school = []
            admitted = []
            personal_website = []
            cell = []
            print("Link:",link)
            driver = webdriver.Chrome()
            def interceptor(request):
                del request.headers['Referer']  # Delete the header first
                request.headers['Referer'] = 'some_referer'

            # Set the interceptor on the driver
            driver.request_interceptor = interceptor

            # All requests will now use 'some_referer' for the referer
            driver.get(link)

            content = driver.page_source
            soup = BeautifulSoup(content)
            cleaner = clean.Cleaner()
            content = cleaner.clean_html(content)
            doc = html.fromstring(content)
            print(soup)
            type(doc)

            type(content)
            
        
            
           
            fo = open('content.txt','w',encoding="utf-8").write(str(content))

            
            '''
            with open('content.txt',encoding="utf-8" ) as f:
                lines = f.readlines()

            print(lines)
            '''
            import re

            textfile = open('content.txt', 'r',encoding="utf-8")
            filetext = textfile.read()
            textfile.close()
            emails = re.findall("[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", filetext)
            print(emails[0])
            
            first_result = (emails[0])
            print(name_final)
            print(first_result)
            final_list.append([i,first_result,query,link])
            stringresult = str(first_result)
            
            if savecount >= 2:
                with open('attorney_list.csv','w', newline='')as f:
                    fields = ['i','first_result','Name','link']
                    writer = csv.writer(f)
                    writer.writerow(fields)
                    writer.writerows(final_list)
                    f.close()
                    savecount = 0
            
            if link_counter == 8 and any(ext in stringresult for ext in blacklist_email): #if no link found by end of loop enters blank into final_list
                final_list.remove([i,first_result,query,link])
                final_list.append([i,[],query,link])
                print('first')
                break
            elif link_counter == 5 and len(first_result) < 1:
                print('second')
                break

            elif len(first_result) < 1 :
                final_list.remove([i,first_result,query,link])
            elif any(ext in stringresult for ext in blacklist_email):
                
                print("Bad email, fetching next link... ",first_result)
                final_list.remove([i,first_result,query,link])
            
                
                
            else:
             print('third')
             break

        else:
            print('fourth')
            continue
            
with open('attorney_list.csv','w', newline='')as f:
    fields = ['i','first_result', 'name','link']
    print(final_list)
    writer = csv.writer(f)
    writer.writerow(fields)
    writer.writerows(final_list)
f.close()
        
print(final_list)





