

from requests_html import HTMLSession
import re
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import csv

data = pd.read_excel(r"C:\Users\darie\Python\email_extraction\scrape imput data\loopnet1test.xlsx")
print (data)
counter = 2
df = pd.DataFrame(data, columns= ['name_email'])

final_list =[]
final = {0: ["query", "emails", "link"]}
wrkbk = openpyxl.load_workbook(r"C:\Users\darie\Python\email_extraction\scrape imput data\loopnet1test.xlsx")
sh = wrkbk.active
for i in range(2,sh.max_row+1): #sh.max_row+1
    print("\n")
    print(i)
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    for link in search(query, tld="co.in", num =1, stop=1, pause =1):  #chage stop= to change how many links are searched
        
        '''

        run loop

        set conditions
        if pass continue
        if it fails clear search loop back try next link





'''


        session = HTMLSession()
        from requests.exceptions import Timeout
        try:
            r = session.get(link, timeout = 10)
        except Timeout:
            print('Timeout has been raised.')
       # r = session.get(link, timeout = 10)
        emails = set()  
      # r.html.render()
        #re.html.render()
        
        print(link)
        try:
            body = r.html.find("body")[0]
        except:
            continue
        original_url = link
       # r=s.get(link,headers=headers)
        soup = BeautifulSoup(r.content, 'html5lib')
        emails = re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", body.text)
        
        loop_result = (emails[:1])
        if [loop_result] != 'support@streeteasy.com' and 'info' not in [loop_result]:
                    first_result = loop_result
        else:
            print(loop_result)
            first_result = ''
        final_list.append([i,query,first_result, link])
        print(first_result)
        loop_result =first_result 
        if len(first_result) < 1:
            final_list.remove([i,query,first_result, link])
            for link in search(query, tld="co.in", num =1, stop=5, pause =5): 
                
                 #chage stop= to change how many links are searched
                session = HTMLSession()
                r = session.get(link)
              
            # r.html.render()
        #re.html.render()
                print(link)
                try:
                    body = r.html.find("body")[0]
                except:
                    continue
                original_url = link
            #   r=s.get(link,headers=headers)
                soup = BeautifulSoup(r.content, 'html5lib')
                emails = re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", body.text)
                loop_result = (emails[:1])
                if loop_result != ['support@streeteasy.com'] and 'info' not in loop_result:
                    first_result = loop_result
                else:
                    print(loop_result)
                print('Trial #2', first_result)
                final_list.append([i,query,first_result, link])
                
                if len(first_result) >= 1:
                    break
            
                else:
                    final_list.remove([i,query,first_result, link])
            if len(first_result) <1:
                final_list.append([i,query,first_result, link])
                

                '''
                for link in search(query, tld="co.in", num =3, stop=3, pause =3): 
                    #chage stop= to change how many links are searched
                    session = HTMLSession()
                    r = session.get(link)
                
        # r.html.render()
            #re.html.render()
                print(link)
                try:
                    body = r.html.find("body")[0]
                except:
                    continue
                original_url = link
            # r=s.get(link,headers=headers)
                soup = BeautifulSoup(r.content, 'html5lib')
                emails = re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", body.text)
                first_result = (emails[:1])
                print('Trial #2', first_result)
                final_list.append([i,query,first_result, link])
            else:
                
            
            
                print(first_result)
        '''
        
        
            
        quotes=[]  # a list to store quotes
   
        table = soup.find('div', attrs = {'id':'email'}) 
        print 
'''
for i in len(final_list):
    for x in final_list[i][2]:

        if 'info' in x or 'streeteasy' in x:
            x = ''
'''

with open('Loopnet_Output_1.csv','w', newline='')as f:
    fields = ['i','Name','first_result','link']
    print(final_list)
    writer = csv.writer(f)
    writer.writerow(fields)
    writer.writerows(final_list)
f.close()


