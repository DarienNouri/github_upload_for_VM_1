import scrapy
from scrapy.spiders import CrawlSpider, Request
from googlesearch import search
import re
from scrapy_selenium import SeleniumRequest
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import openpyxl

url_bank = []


data = pd.read_excel(r"C:\Users\darie\OneDrive\Documents\1NYU\python\For Internship Web Scraping\python web scrape.xlsx")
print (data)

df = pd.DataFrame(data, columns= ['name_email'])
print(df)

final = {}
wrkbk = openpyxl.load_workbook(r"C:\Users\darie\OneDrive\Documents\1NYU\python\For Internship Web Scraping\python web scrape.xlsx")
sh = wrkbk.active
for i in range(2, sh.max_row+1): #sh.max_row+1
    print("\n")
    print(i)
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    for url in search(query, tld="co.in", num =1, stop=1, pause =2):
        #r = requests.get(url)
        print(url)
        url_bank.append(url)
    
    class email_extractor(CrawlSpider):
     
    # adjusting parameters
        name = 'email_ex'
 
        def __init__(self, *args, **kwargs):
            super(email_extractor, self).__init__(*args, **kwargs)
            self.email_list = []
            self.query = " 'market places'.gmail.com "
            
        
        def start_requests(self):
            for results in search(self.query, num=10, stop=None, pause=2):
                yield SeleniumRequest(
                    url=results,
                    callback=self.parse,
                    wait_until=EC.presence_of_element_located(
                        (By.TAG_NAME, "html")),
                    dont_filter=True
                )
        def parse(self, response):
            EMAIL_REGEX = r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+'
            emails = re.finditer(EMAIL_REGEX, str(response.text))
            for email in emails:
                self.email_list.append(email.group())
    
            for email in set(self.email_list):
                yield{
                    "emails": email
                }
    
        
from googlesearch import search
import pandas as pd 
import openpyxl
import requests
from bs4 import BeautifulSoup
import re
from scrapy_selenium import SeleniumRequest
from selenium.webdriver.common.by import By

import urllib.parse
from collections import deque
import files
import pandas as pd
from requests_html import HTMLSession






data = pd.read_excel(r"C:\Users\darie\OneDrive\Documents\1NYU\python\For Internship Web Scraping\python web scrape.xlsx")
print (data)

df = pd.DataFrame(data, columns= ['name_email'])

final_list =[]
final = {0: ["query", "emails", "link"]}
wrkbk = openpyxl.load_workbook(r"C:\Users\darie\OneDrive\Documents\1NYU\python\For Internship Web Scraping\python web scrape.xlsx")
sh = wrkbk.active
for i in range(75,85): #sh.max_row+1
    print("\n")
    print(i)
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    for link in search(query, tld="co.in", num =1, stop=3, pause =3):
        r = session.get(link)
        emails = set()  
      # r.html.render()
        original_url = link
       # r=s.get(link,headers=headers)
        soup = BeautifulSoup(r.content, 'html5lib')
   
        quotes=[]  # a list to store quotes
   
        table = soup.find('div', attrs = {'id':'email'}) 
        print 












'''


file_object = open('testemails.txt', 'w')
data = pd.read_excel(r"C:\Users\darie\OneDrive\Documents\1NYU\python\For Internship Web Scraping\python web scrape.xlsx")
print (data)

df = pd.DataFrame(data, columns= ['name_email'])
print(df)

final = []
wrkbk = openpyxl.load_workbook(r"C:\Users\darie\OneDrive\Documents\1NYU\python\For Internship Web Scraping\python web scrape.xlsx")
sh = wrkbk.active
for i in range(2, sh.max_row+1): #sh.max_row+1
    print("\n")
    print(i)
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    for url in search(query, tld="co.in", num =1, stop=1, pause =2):
        #r = requests.get(url)
        print(url)
        
        EMAIL_REGEX = r"""(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])"""


        session = HTMLSession()
        
        try:
            r = session.get(url)
        except:
            r.html.render()
        
        body = r.html.find("body")[0]
        emails = re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", body.text)
        final.append([emails])
        for index,email in enumerate(emails):
            print(index+1, "---->", email)
        
print(emails)  


        emails = re.findall.EMAIL_REGEX, r.html.raw_html.decode()):
            print(re_match.group())


        '''