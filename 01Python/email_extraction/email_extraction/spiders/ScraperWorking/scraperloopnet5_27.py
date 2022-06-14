
from requests_html import HTMLSession
import re
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import csv
import time

data = pd.read_excel(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\loopnet1test.xlsx")
print (data)
counter = 2
df = pd.DataFrame(data, columns= ['name_email'])
savecount = 0
saveloop = 0


final_list =[]
final = {0: ["emails", "query", "link"]}
wrkbk = openpyxl.load_workbook(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\loopnet1test.xlsx")
sh = wrkbk.active
for i in range(2,sh.max_row+1): #sh.max_row+1
    print("\n")
    print(i)
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    link_counter = 0
    savecount+=1
    for link in search(query, tld="co.in", num =1, stop=5, pause =5):  #chage stop= to change how many links are searched
        link_counter+=1
        print(query)
        print(link)
        session = HTMLSession()
        
        req_headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.8',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36'
        }
        
        try:
            r = session.get(link,  headers=req_headers)
        except:
            print("badlink")
            continue
        

        if r.status_code == 429:
             time.sleep(10)
        time.sleep(.8)
        emails = set()  
        #r.html.render()
        #re.html.render()
        blacklist_link = ["rocket","linkedin","streeteasy","crexi","zoom","loopnet","zillow"]
        blacklist_email = ["info", "customer", "streeteasy", "help","edu",".org","Info","_","fake","foo", "help","contact","jane","doe",".ru",".io"]
        try:
            body = r.html.find("body")[0]
        except:
            if link_counter == 5:
                final_list.append([i,[],query, link])
            print('veryfirst')
            '''
            try:
                final_list.remove([i,query,first_result, link])
            except:
                final_list.append([i,query,[], link])
            '''


            continue
                
        original_url = link
       # r=s.get(link,headers=headers)
        
        
        if any(ext not in link for ext in blacklist_link): 
            print("Link:",link)
            
            soup = BeautifulSoup(r.content, 'html5lib')
            emails = re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", body.text)
            first_result = (emails[:1])
            print(name_final)
            print(first_result)
            final_list.append([i,first_result,query,link])
            stringresult = str(first_result)
            
            if savecount >= 2:
                with open('exportformatingtest.csv','w', newline='')as f:
                    fields = ['i','first_result','Name','link']
                    writer = csv.writer(f)
                    writer.writerow(fields)
                    writer.writerows(final_list)
                    f.close()
                    savecount = 0
            
            if link_counter == 5 and any(ext in stringresult for ext in blacklist_email): #if no link found by end of loop enters blank into final_list
                final_list.remove([i,first_result,query,link])
                final_list.append([i,[],query,link])
                print('first')
                break
            elif link_counter == 5 and len(first_result) < 1:
                print('second')
                break

            elif len(first_result) < 1 :
                final_list.remove([i,first_result,query,link])
            elif any(ext in stringresult for ext in blacklist_email):
                
                print("Bad email, fetching next link... ",first_result)
                final_list.remove([i,first_result,query,link])
            
                
                
            else:
             print('third')
             break

        else:
            print('fourth')
            continue
            
with open('exportformatingtest.csv','w', newline='')as f:
    fields = ['i','first_result', 'name','link']
    print(final_list)
    writer = csv.writer(f)
    writer.writerow(fields)
    writer.writerows(final_list)
f.close()
        
print(final_list)

