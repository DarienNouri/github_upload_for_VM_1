
from requests_html import HTMLSession
import re
from bs4 import BeautifulSoup
import pandas as pd
import csv
import time

link = 'https://www.floridabar.org/directories/find-mbr/profile/?num=561347'

from bs4 import BeautifulSoup
import urllib3
from seleniumwire import webdriver

moon_url = 'https://www.floridabar.org/directories/find-mbr/profile/?num=564737'


import lxml.html as html
import lxml.html.clean as clean
from seleniumwire import webdriver
import requests
import lxml
from lxml import etree
import xml.etree.ElementTree as ET 
from googlesearch import search
from pyvirtualdisplay import Display
from selenium.webdriver import DesiredCapabilities
import pandas as pd
from selenium.webdriver.common.by import By
import openpyxl


options = webdriver.ChromeOptions()

data = pd.read_excel(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\lawyer_input.xlsx")
print (data)
counter = 2
df = pd.DataFrame(data, columns= ['name_email'])
savecount = 0
saveloop = 0

end_list = [] #actual final list
final_list =[]
final = {0: ["emails", "query", "link"]}
wrkbk = openpyxl.load_workbook(r"C:\Users\darie\OneDrive\Documents\01Python\email_extraction\scrape imput data\lawyer_input.xlsx")
sh = wrkbk.active
for i in range(4,sh.max_row+1): #sh.max_row+1
    print("\n")
    
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
    try:
        from googlesearch import search
    except ImportError:
        print('doesnt work')
    name_final = str(cell_obj.value)
    query = str(cell_obj.value)
    link_counter = 0
    savecount+=1
    for link in search(query, tld="co.in", num =1, stop=9, pause =9):  #chage stop= to change how many links are searched
        time.sleep(1)
        
        link_counter+=1
        
        session = HTMLSession()
        
        req_headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.8',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/4.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36'
        }
        
        try:
            r = session.get(link,  headers=req_headers)
        except:
            print("badlink")
            continue
        if r.status_code == 429:
             time.sleep(4)
        
    
      
        blacklist_link = ["rocket","linkedin","streeteasy","crexi","zoom","loopnet","zillow"]
        blacklist_email = ["info", "customer", "streeteasy", "help","edu",".org","Info","_","fake","foo", "help","contact","jane","doe",".ru",".io"]
        whitelist_url = ['floridabar']
        


        if any(ext in link for ext in whitelist_url): 
            fin_list = []
            address = []
            phone = []
            area = []
            firm = []
            law_school = []
            admitted = []
            personal_website = []
            cell = []
            
            
            options = webdriver.ChromeOptions()

            def interceptor(request):
                del request.headers['Referer']  # Delete the header first
                request.headers['Referer'] = 'some_referer'
                options.add_experimental_option('excludeSwitches', ['enable-logging'])
            driver = webdriver.Chrome(options=options)
            # Set the interceptor on the driver
            driver.request_interceptor = interceptor
            # All requests will now use 'some_referer' for the referer
            try:
                driver.get(link)
            except:
                break
            content = driver.page_source
            cleaner = clean.Cleaner()
            content = cleaner.clean_html(content)
        
            content1 = content
            html = content1
            soup = BeautifulSoup(html,features="lxml")

            a_tags = soup.select("a[href^=tel]")
            title = soup.find("h4", class_="col-xs-12 col-sm-8")
            from pprint import pprint as pp

            dom = etree.HTML(str(soup))
            p = dom.xpath("//p")[0]
            print()
        
            profile = driver.find_element(By.ID, 'mProfile').text
            
            df = pd.DataFrame(columns= ['city',
                                    'circut',
                                    'date_licenced',
                                    '10-yr',
                                    'school',
                                    'practice',
                                    'firm',
                                    'size_firm',
                                    'position',
                                    ])

            row = 2

            dict= {1:['city','date_licenced','school',
                    'practice','firm','size_firm','fd']}  
            df = pd.DataFrame(dict)

            name = soup.find("h1", class_="full").text
            addy = driver.find_elements(By.CLASS_NAME,'col-sm-8')
            email = driver.find_elements(By.CLASS_NAME, 'icon-email')[0].text
            bar_url = driver.find_elements(By.CLASS_NAME, 'col-sm-9')
            addy_list = []
            fin_list = []



            for x in bar_url:  # remove unwanted
                if 'Committee' or 'www' in x:
                    bar_url.remove(x)
                    
            print('start_bar')
            for i in bar_url:
                print(i.text)
            unsliced = addy[0].text
            
            for i in unsliced:
                if 'www.' in i:
                    print(i)
                    unsliced.pop(i)
            sliced1 = unsliced.split('\n')
            print(sliced1)
            
            print(sliced1)


##set variables here
            for row in sliced1:
                if 'Palm Beach' in row:
                    city = row
                elif row[0].isdigit:
                    admitted = row
                elif 'Real Property' or 'Probate' in row:
                    specialty = row
                

            
            print()
            print()
            unsliced = addy[1].text
            sliced2 = unsliced.split('\n')
            print(sliced2)
            break
            '''
            addy_list.append(['Address',addy[0].text,addy[1].text])
            fin_list.append(['Name', name])
            fin_list.append(['Address',addy_list])
            bar_profile = bar_url[0].text 
            city = bar_url[1].text
            year_bar = bar_url[2].text
            school = bar_url[3].text
            area =  bar_url[4].text
            
            fin_list.append(['Bar Profile',bar_profile])
            fin_list.append(['City', city])
            fin_list.append(['Year Bar', year_bar])
            fin_list.append(['School', school])
           
            for x in bar_url: # apend bar_url
                #print((x)[1].text)
                break
            print(fin_list)
            end_list.append([fin_list])
                
            if savecount >=2:
                
                with open('attorney_list.csv','w', newline='')as f:
                    writer = csv.writer(f)
                    writer.writerows(end_list)
                f.close()
                savecount = 0

            print(end_list)
            print()
            break
                                
          
                
           
    
        
print(final_list)




'''
